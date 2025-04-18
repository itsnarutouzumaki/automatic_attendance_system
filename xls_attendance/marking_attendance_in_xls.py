import openpyxl

from datetime import datetime
import logging
logging.getLogger('comtypes').setLevel(logging.WARNING)


def mark_attendance(course_id, registration_numbers):
    filename = f"attendance_record/{course_id}.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        registration_column = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == "Registration Number":
                registration_column = col[0].column
                break

        if not registration_column:
            print("Error: 'Registration Number' column not found in the file.")
            return []
        
        attendance_column = sheet.max_column + 1
        sheet.cell(row=1, column=attendance_column, value=datetime.now().strftime("%d-%m-%Y"))

        successfully_marked = []

        for reg_number in registration_numbers:
            reg_number = reg_number.lower()
            found = False

            for row in sheet.iter_rows(min_row=2, max_col=registration_column, max_row=sheet.max_row):
                if row[0].value and row[0].value.lower() == reg_number:
                    sheet.cell(row=row[0].row, column=attendance_column, value="P")
                    successfully_marked.append(reg_number)
                    found = True
                    break

            if not found:
                print(f"Warning: Registration number '{reg_number}' not found in the sheet.")

        for row in sheet.iter_rows(min_row=2, max_col=registration_column, max_row=sheet.max_row):
            if row[0].value and not sheet.cell(row=row[0].row, column=attendance_column).value:
                sheet.cell(row=row[0].row, column=attendance_column, value="A")

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(filename)
        print(f"Attendance marked successfully in '{filename}'.")
        return successfully_marked

    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return []
