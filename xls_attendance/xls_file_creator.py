import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def create_attendance_file(course_code):
    folder_name = "attendance_record"
    os.makedirs(folder_name, exist_ok=True)

    filename = os.path.join(folder_name, f"{course_code}.xlsx")

    if os.path.exists(filename):
        print(f"Attendance file '{filename}' already exists. Skipping creation.")
        return None  # File already exists, so return None

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Registration Number'

    column_letter = get_column_letter(1)
    sheet.column_dimensions[column_letter].width = 20
    sheet['A1'].alignment = Alignment(horizontal='center')

    workbook.save(filename)
    print(f"Attendance file '{filename}' created successfully.")
    return filename

def add_registration_numbers(filename, registration_numbers):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for reg_number in registration_numbers:
        if reg_number.strip():
            sheet.append([reg_number.strip().lower()])

    workbook.save(filename)
    print(f"Registration numbers saved to '{filename}'.")

def xls_file_creator(course_code, registration_numbers):
    if not course_code:
        print("Course code cannot be empty. Exiting.")
        return

    if not registration_numbers:
        print("Registration numbers list is empty. Exiting.")
        return

    filename = create_attendance_file(course_code)

    if filename:  # Only add registration numbers if file was newly created
        add_registration_numbers(filename, registration_numbers)
    else:
        print("No registration numbers were added as the file already exists.")

